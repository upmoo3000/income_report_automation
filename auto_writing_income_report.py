'''
[필요 파일 목록]
- transaction_history = 거래내역조회서
- students_list = 입사자 명단
- income_report = 수입보고서

[사용 방법]
1. 작성할 기간의 거래내역조회서 다운 후 다른 이름으로 저장하여 xlsx 확장자로 저장
2. 파일명 '[월]_transaction_history'로 변경
3. auto_writing_income_report.py 내 18행의 불러올 파일명을 방금 다운 받은 거래내역조회서의 파일명과 똑같이 변경
4. ctrl + alt + N 하면 파이썬 코드가 실행되어 수입결의서에 학생 이름 자동 입력
5. 수입결의서 열어 학생 및 기타 서식 정리(노란색: 이름만 불일치, 빨간색: 이름&입금자 둘 다 불일치)
'''
import pandas as pd
import openpyxl as op
import re
import gspread as gs

# 환경변수
from dotenv import load_dotenv
import os

load_dotenv()
googleApiKey = os.environ.get("GOOGLE_API_KEY")
studentsListSheetURL = os.environ.get("STUDENTS_LIST_SHEET_URL")
activatedSheetName = os.environ.get("ACTIVATED_SHEET_NAME")
transactionFile = os.environ.get("TRANSACTION_FILE_NAME")
incomeReportFile = os.environ.get("INCOME_REPORT_FILE_NAME")

# 엑셀 파일 불러오기
transactions = pd.read_excel(transactionFile)
wb = op.load_workbook(incomeReportFile) 
ws = wb.active
# Google Sheets
gc = gs.service_account(filename=googleApiKey)
doc = gc.open_by_url(studentsListSheetURL)
worksheet = doc.worksheet(activatedSheetName)

# 스타일 설정
styles = op.styles   

# 입금내역 변수 저장 및 입사자명단에서 학생명 추출하여 배열화
depositors = transactions['내역']
studentNamesFromStudentsList = worksheet.col_values(4)
depositorNamesFromStudentsList = worksheet.col_values(15)
incomeReportNameCol = ws['B']
onlyDepositorNames = []       

# 입금내역에서 숫자 제거하여 문자만 추출
def extractNonNumeric(val):
    nonNumeric = re.sub(r'\d+', '', val).strip()
    return nonNumeric

# 입금자명 배열 전체 탐색하며 숫자 제거하여 onlyDepositorNames append
for val in depositors:
    if type(val) is float:
        continue
    onlyDepositorNames.append(extractNonNumeric(val))

# 입금자명행에서 입금자명 유무 확인
def searchStudentIdx(name):
    for i, val in enumerate(studentNamesFromStudentsList, 0):
        studentName = str(val)
        if studentName == 'nan' or name not in studentName:
            continue
        if name == studentName:
            return i

# 입금자명행에서 입금자명 유무 확인
def searchDepositorIdx(name):
    for i, val in enumerate(depositorNamesFromStudentsList, 0):
        depositorName = str(val)
        if depositorName == 'nan' or name not in depositorName:
            continue
        if name in depositorName:
            return i

# 결괏값 삽입
for i, name in enumerate(onlyDepositorNames, 0):
    # 입금완료열 탐색
    studentIdx = searchStudentIdx(name) # 입사자명단 이름열에서 입금자명 검색
    if studentIdx: 
        ws['B'+ str(i + 2)].value = studentNamesFromStudentsList[studentIdx]
        continue
    depositorIdx = searchDepositorIdx(name) # 이름열에서 입금자명 검색 안 될 시 입금자명열에서 입금자명 검색 후 삽입
    if depositorIdx:
        ws['B'+ str(i + 2)].value = studentNamesFromStudentsList[depositorIdx] + depositorNamesFromStudentsList[depositorIdx]
        ws['B'+ str(i + 2)].fill = styles.PatternFill(fill_type ='solid', fgColor = styles.Color('FFFF00'))
    else: # 이름열 및 입금자명열 둘 다 검색 안 될 시 확인 필요 문구 삽입&빨간색칠
        ws['B'+ str(i + 2)].value = name + ' 확인 필요'
        ws['B'+ str(i + 2)].fill = styles.PatternFill(fill_type ='solid', fgColor = styles.Color('FFFF0000'))
                
# 수입결의서 저장
wb.save('income_report.xlsx')
